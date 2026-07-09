from flask import Flask, render_template, request, jsonify, send_file
import requests
import json
import io
import re
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# --- Carrega ClassTrib da planilha oficial LC 214/2025 ---
_CLASSTRIB_FILE = os.path.join(os.path.dirname(__file__), 'cClassTrib_2025-12-12.xlsx')

def _load_classtrib():
    if not os.path.exists(_CLASSTRIB_FILE):
        return [], []
    wb = load_workbook(_CLASSTRIB_FILE, data_only=True)

    # Aba cClassTrib
    ws_class = wb['cClass 12-12-2025']
    rows = list(ws_class.iter_rows(values_only=True))
    non_empty = [r for r in rows if any(c is not None for c in r)]
    header = non_empty[0]

    def _fmt_date(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime('%d/%m/%Y')
        return str(val)

    classtrib = []
    for row in non_empty[1:]:
        cst = str(row[0]).zfill(3) if row[0] is not None else ''
        code = str(row[2]).zfill(6) if row[2] is not None else ''
        if not code or code == '000000':
            continue
        item = {
            'cst_ibs_cbs': cst,
            'desc_cst': str(row[1]) if row[1] else '',
            'cClassTrib': code,
            'nome': str(row[3]) if row[3] else '',
            'descricao': str(row[4]) if row[4] else '',
            'lc_redacao': str(row[5]) if row[5] else '',
            'lc_214': str(row[6]) if row[6] else '',
            'tipo_aliquota': str(row[7]) if row[7] else '',
            'pRedIBS': float(row[8]) if row[8] is not None else 0,
            'pRedCBS': float(row[9]) if row[9] is not None else 0,
            'ind_trib_regular': int(row[10]) if row[10] is not None else 0,
            'ind_cred_pres': int(row[11]) if row[11] is not None else 0,
            'ind_mono_padrao': int(row[12]) if row[12] is not None else 0,
            'ind_mono_reten': int(row[13]) if row[13] is not None else 0,
            'ind_mono_ret': int(row[14]) if row[14] is not None else 0,
            'ind_mono_dif': int(row[15]) if row[15] is not None else 0,
            'ind_estorno_cred': int(row[16]) if row[16] is not None else 0,
            'd_ini_vig': _fmt_date(row[17]),
            'd_fim_vig': _fmt_date(row[18]),
            'd_atualizacao': _fmt_date(row[19]),
            'ind_nfe': int(row[21]) if row[21] is not None else 0,
            'ind_nfce': int(row[22]) if row[22] is not None else 0,
            'ind_cte': int(row[23]) if row[23] is not None else 0,
            'anexo': str(row[35]) if row[35] else '',
            'link': str(row[36]) if row[36] else '',
        }
        classtrib.append(item)

    # Aba CST
    ws_cst = wb['CST 12-12-2025']
    rows2 = list(ws_cst.iter_rows(values_only=True))
    non2 = [r for r in rows2 if any(c is not None for c in r)]
    cst_list = []
    for row in non2[1:]:
        if row[0] is not None and isinstance(row[0], (int, float)) and row[1]:
            cst_list.append({
                'cst': str(int(row[0])).zfill(3),
                'descricao': str(row[1]),
                'ind_ibs_cbs': int(row[2]) if row[2] is not None else 0,
                'ind_mono': int(row[3]) if row[3] is not None else 0,
                'ind_red': int(row[4]) if row[4] is not None else 0,
                'ind_dif': int(row[5]) if row[5] is not None else 0,
                'ind_transf_cred': int(row[6]) if row[6] is not None else 0,
            })
    return classtrib, cst_list

CLASSTRIB_DATA, CST_IBS_CBS_DATA = _load_classtrib()

# --- Tabela de produtos monofásicos ---
MONOFASICOS = [
    # COMBUSTÍVEIS
    {"ncm": "27101259", "descricao": "Gasolinas, exceto de aviação", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101219", "descricao": "Gasolina de aviação", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101921", "descricao": "Óleo diesel", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101922", "descricao": "Óleo diesel marítimo", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101931", "descricao": "Querosene de aviação (QAV)", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101939", "descricao": "Outros querosenes", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27101100", "descricao": "Óleos leves e preparações", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27111100", "descricao": "Gás natural liquefeito (GNL)", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27112100", "descricao": "Gás natural em estado gasoso", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27111910", "descricao": "GLP - Gás liquefeito de petróleo", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27111200", "descricao": "Propano liquefeito", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27111300", "descricao": "Butano liquefeito", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "27112900", "descricao": "Outros gases de petróleo liquefeitos", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "22071000", "descricao": "Álcool etílico não desnaturado — uso predominante como combustível (AEHC); o mesmo NCM também cobre álcool não desnaturado destinado a bebidas, cuja tributação segue regime distinto (a classificação NCM não distingue por destinação — verificar o uso real do produto)", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º (uso combustível)"},
    {"ncm": "22072000", "descricao": "Álcool etílico desnaturado (AEAC)", "categoria": "Combustíveis", "referencia": "Lei 9.718/98 art. 4º"},
    {"ncm": "38260000", "descricao": "Biodiesel e suas misturas", "categoria": "Combustíveis", "referencia": "Lei 11.116/2005 art. 1º"},

    # MEDICAMENTOS
    {"ncm": "30039011", "descricao": "Medicamentos com penicilina não dosados", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30039019", "descricao": "Outros antibióticos não dosados", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30039099", "descricao": "Outros medicamentos não dosados", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30041011", "descricao": "Medicamentos com penicilina para uso humano", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30041019", "descricao": "Outros med. com antibióticos uso humano", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30042011", "descricao": "Antibióticos uso humano (outros grupos)", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30043100", "descricao": "Medicamentos com insulina", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30043900", "descricao": "Outros medicamentos com hormônios", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30044200", "descricao": "Medicamentos com cloridrato de efedrina", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30044900", "descricao": "Outros medicamentos com alcaloides", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30045011", "descricao": "Medicamentos com vitamina A", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30049019", "descricao": "Outros medicamentos para uso humano", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "30049099", "descricao": "Outros medicamentos (uso humano)", "categoria": "Medicamentos", "referencia": "Lei 10.147/2000 art. 1º"},

    # PERFUMES E COSMÉTICOS
    {"ncm": "33030010", "descricao": "Perfumes (extratos)", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33030020", "descricao": "Águas-de-colônia", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33041000", "descricao": "Produtos de maquilagem labial", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33042010", "descricao": "Sombras, delineadores, sobrancelhas", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33042090", "descricao": "Outros produtos de maquilagem olhos", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33043000", "descricao": "Preparações para manicure/pedicure", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33049100", "descricao": "Pós de maquilagem (incluindo compactos)", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33049900", "descricao": "Outros produtos de beleza/maquilagem", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33052000", "descricao": "Preparações para ondulação/alisamento", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33053000", "descricao": "Laquês para o cabelo", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33059000", "descricao": "Outras preparações capilares", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33071000", "descricao": "Preparações para barbear", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33072000", "descricao": "Desodorantes corporais e antiperspirantes", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33074100", "descricao": "Agarbatti e preparações odoríferas p/ queima", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "33079000", "descricao": "Outros produtos toucador/cosmética", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º"},
    {"ncm": "34011190", "descricao": "Sabões de toucador, em barras/pedaços/figuras moldadas (verificar exceção 'Ex 01' da TIPI antes de aplicar)", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º, I, 'b'"},
    {"ncm": "34012010", "descricao": "Sabões de toucador sob outras formas", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º, I, 'b'"},
    {"ncm": "96032100", "descricao": "Escovas de dentes", "categoria": "Perfumes e Cosméticos", "referencia": "Lei 10.147/2000 art. 1º, I, 'b'"},

    # VEÍCULOS
    {"ncm": "87011000", "descricao": "Motocultores", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87012000", "descricao": "Tratores rodoviários para semirreboque", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87013700", "descricao": "Tratores agrícolas de potência até 18kW", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87021100", "descricao": "Ônibus/microônibus motor diesel/semidiesel", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87031000", "descricao": "Veículos para neve; veículos de golfe", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87032110", "descricao": "Automóveis até 1000cc gasolina", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87032190", "descricao": "Outros automóveis até 1000cc gasolina", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87032210", "descricao": "Automóveis 1000-1500cc gasolina", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87032290", "descricao": "Outros automóveis 1000-1500cc gasolina", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87032310", "descricao": "Automóveis 1500-3000cc gasolina", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87033110", "descricao": "Automóveis até 1500cc diesel", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87033210", "descricao": "Automóveis 1500-2500cc diesel", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87041000", "descricao": "Veículos para transporte de mercadorias dumper", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87042110", "descricao": "Caminhões diesel PBT <= 5t", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87042190", "descricao": "Outros caminhões diesel PBT <= 5t", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87042210", "descricao": "Caminhões diesel PBT 5-20t", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87042290", "descricao": "Outros caminhões diesel PBT 5-20t", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87051000", "descricao": "Caminhões guindaste", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87060010", "descricao": "Chassis com motor para ônibus", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87060090", "descricao": "Outros chassis com motor p/ veículos", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87111000", "descricao": "Motocicletas motor pistão até 50cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87112010", "descricao": "Motocicletas 50-125cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87112090", "descricao": "Outras motocicletas 50-125cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87113000", "descricao": "Motocicletas 125-250cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87114000", "descricao": "Motocicletas 250-500cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87115000", "descricao": "Motocicletas 500-800cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "87119010", "descricao": "Motocicletas acima 800cc", "categoria": "Veículos", "referencia": "Lei 10.485/2002 art. 1º"},

    # MÁQUINAS E EQUIPAMENTOS
    {"ncm": "84071000", "descricao": "Motores de aviação a explosão", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84073100", "descricao": "Motores explosão veículos cap. <= 50cc", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84073200", "descricao": "Motores explosão veículos 50-250cc", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84073300", "descricao": "Motores explosão veículos 250-1000cc", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84073400", "descricao": "Motores explosão veículos > 1000cc", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84081000", "descricao": "Motores diesel para embarcações", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},
    {"ncm": "84082000", "descricao": "Motores diesel para veículos", "categoria": "Máquinas e Equipamentos", "referencia": "Lei 10.485/2002 art. 1º"},

    # AUTOPEÇAS
    {"ncm": "84091000", "descricao": "Partes para motores de aviação", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84099100", "descricao": "Partes para motores explosão", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84099900", "descricao": "Partes para motores diesel", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84131100", "descricao": "Bombas de combustível para motores", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84131900", "descricao": "Outras bombas combustível/lubrificante", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84141000", "descricao": "Bombas de vácuo para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84212300", "descricao": "Filtros de óleo para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)"},
    {"ncm": "84229090", "descricao": "Partes de máquinas de lavar veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84831010", "descricao": "Virabrequins para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84833010", "descricao": "Caixas de marchas para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84834010", "descricao": "Engrenagens e rodas de fricção p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "84835010", "descricao": "Volantes e polias p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "85122010", "descricao": "Aparelhos de sinalização de veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "85123010", "descricao": "Aparelhos de sinalização acústica veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "85124010", "descricao": "Limpadores de para-brisas", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "85131010", "descricao": "Lanternas de mão para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87082110", "descricao": "Cintos de segurança para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87082910", "descricao": "Outras partes e acessórios de carroceria", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87083110", "descricao": "Freios e servo-freios p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87083910", "descricao": "Outras partes de freios p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87084010", "descricao": "Caixas de marchas p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87085010", "descricao": "Eixos motrizes e não motrizes p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87086010", "descricao": "Rodas e respectivos acessórios p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87087010", "descricao": "Radiadores p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87088010", "descricao": "Silenciosos e tubos de escape p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87089110", "descricao": "Radiadores de ar condicionado p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87089210", "descricao": "Volantes, colunas e caixas de direção", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "87089910", "descricao": "Outras partes e acessórios p/ veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 3º"},
    {"ncm": "40111000", "descricao": "Pneus novos para automóveis", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40112000", "descricao": "Pneus novos para ônibus e caminhões", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40113000", "descricao": "Pneus novos para aeronaves", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40114000", "descricao": "Pneus novos para motocicletas", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40115000", "descricao": "Pneus novos para bicicletas", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40117090", "descricao": "Pneus novos para tratores agrícolas/florestais", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40118090", "descricao": "Pneus novos para máquinas de construção/terraplanagem", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40119090", "descricao": "Outros pneus novos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40121100", "descricao": "Pneus recauchutados para automóveis", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40121200", "descricao": "Pneus recauchutados para ônibus e caminhões", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40121300", "descricao": "Pneus recauchutados para aeronaves", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40121900", "descricao": "Outros pneus recauchutados", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40122000", "descricao": "Pneus usados", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40129090", "descricao": "Outras partes de pneus (protetores, flaps etc.)", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40131090", "descricao": "Câmaras de ar para automóveis/ônibus/caminhões", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40132000", "descricao": "Câmaras de ar para bicicletas", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},
    {"ncm": "40139000", "descricao": "Outras câmaras de ar", "categoria": "Autopeças", "referencia": "Lei 10.485/2002 art. 5º"},

    # BEBIDAS FRIAS (água, refrigerante, cerveja — Lei 13.097/2015, art. 14)
    {"ncm": "22011000", "descricao": "Água mineral natural", "categoria": "Bebidas Frias", "referencia": "Lei 13.097/2015, art. 14"},
    {"ncm": "22019000", "descricao": "Outras águas (incluindo gelo)", "categoria": "Bebidas Frias", "referencia": "Lei 13.097/2015, art. 14"},
    {"ncm": "22021000", "descricao": "Água, incl. mineral e gaseificada c/ açúcar", "categoria": "Bebidas Frias", "referencia": "Lei 13.097/2015, art. 14"},
    {"ncm": "22029000", "descricao": "Outras bebidas não alcoólicas", "categoria": "Bebidas Frias", "referencia": "Lei 13.097/2015, art. 14"},
    {"ncm": "22030000", "descricao": "Cervejas de malte", "categoria": "Bebidas Frias", "referencia": "Lei 13.097/2015, art. 14"},

    # DEMAIS BEBIDAS (vinhos, fermentados, destilados — regime distinto, não abrangido pela Lei 13.097/2015)
    {"ncm": "22041000", "descricao": "Vinhos espumantes e champanha", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22042100", "descricao": "Vinhos em recipientes até 2L", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22042900", "descricao": "Outros vinhos em recipientes > 2L", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22043000", "descricao": "Outros mostos de uva", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22060000", "descricao": "Outras bebidas fermentadas (sidra, perada)", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22082000", "descricao": "Aguardente de vinho ou de bagaço de uva", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22083000", "descricao": "Uísques", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22084000", "descricao": "Rum e outras aguardentes de cana-de-açúcar", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22085000", "descricao": "Gim e genebra", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22086000", "descricao": "Vodca", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22087000", "descricao": "Licores", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
    {"ncm": "22089000", "descricao": "Outras bebidas espirituosas", "categoria": "Bebidas", "referencia": "Lei 9.718/98 (regime especial)"},
]

# --- Regras de monofásico por capítulo/posição NCM ---
# Complementam a lista MONOFASICOS acima. Aplicadas somente quando o NCM
# não está listado individualmente (para não duplicar a regra de NCM específico).
#
# Cap. 30: a Lei 10.147/2000, art. 1º, I, "a" só atinge as posições 30.01,
# 30.03 (exceto NCM 3003.90.56) e 30.04 (exceto NCM 3004.90.46), além de
# subitens pontuais de 30.02 (soros/vacinas) e 3006.30 (reagentes/contrastes
# de diagnóstico) que não são modelados aqui por exigirem mapeamento preciso
# ao NCM8 vigente — tratar esses casos manualmente. NÃO inclui 30.05 (gazes,
# ataduras) nem o restante de 30.06 (cimentos dentários, kits de primeiros
# socorros etc.), que seguem regime normal.
#
# Cap. 33: a mesma lei, art. 1º, I, "b", cobre 33.03 a 33.07 exceto o NCM
# 3305.10.00 (xampu, expressamente excluído).
_MONO_CHAPTER_RULES = [
    {
        "chapter": "30",
        "headings": {"3001", "3003", "3004"},
        "excecoes": {"30039056", "30049046"},
        "categoria": "Medicamentos",
        "referencia": "Lei 10.147/2000 art. 1º, I, 'a'",
        "regra_identificacao": "Monofásico por posição NCM — 30.01, 30.03 e 30.04 (Lei 10.147/2000)",
        "descricao_padrao": "Medicamento — posições 30.01, 30.03 ou 30.04",
    },
    {
        # Posição 33.06 (higiene bucal/dentária) removida desta regra: a
        # Lei 10.925/2004 (Cesta Básica, inciso XXVII) zerou o PIS/COFINS
        # da posição inteira, sem ressalva de "Ex" — substitui o regime
        # monofásico anterior por completo para esses produtos.
        "chapter": "33",
        "headings": {"3303", "3304", "3305", "3307"},
        "excecoes": {"33051000"},
        "categoria": "Perfumes e Cosméticos",
        "referencia": "Lei 10.147/2000 art. 1º, I, 'b'",
        "regra_identificacao": "Monofásico por posição NCM — 33.03, 33.04, 33.05 (exceto 3305.10.00) e 33.07 (Lei 10.147/2000)",
        "descricao_padrao": "Perfume/Cosmético/Higiene — posições 33.03, 33.04, 33.05 e 33.07",
    },
    {
        "chapter": "22",
        "headings": {"2201", "2202", "2203"},
        "categoria": "Bebidas Frias",
        "referencia": "Lei 13.097/2015, art. 14",
        "regra_identificacao": "Monofásico por posição NCM — 22.01 a 22.03 'bebidas frias' (Lei 13.097/2015)",
        "descricao_padrao": "Bebida fria (água/refrigerante/cerveja) — posições 22.01 a 22.03",
    },
]

_AVISO_AUTOPROPULSADO = (
    "Aplica-se apenas a produtos autopropulsados (§1º do art. 1º da Lei "
    "10.485/2002). Confirme se este equipamento é autopropulsado antes de "
    "aplicar o benefício."
)

_AVISO_ANEXO_II = (
    "Este NCM está sujeito à tributação monofásica APENAS quando destinado "
    "aos produtos específicos listados no Anexo II da Lei 10.485/2002. "
    "Confirme a destinação/aplicação real do produto antes de aplicar o "
    "benefício."
)

_MONO_HEADING_RULES = [
    {
        "headings": {"8711", "8712", "8713", "8714"},
        "categoria": "Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por posição NCM — 8711 a 8714 (Lei 10.485/2002)",
        "descricao_padrao": "Motocicleta/Veículo",
    },
    # --- Máquinas e veículos, Lei 10.485/2002 art. 1º — aplicação direta ---
    {
        "headings": {"7309"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por posição NCM — 73.09 (Lei 10.485/2002)",
        "descricao_padrao": "Reservatório, tanque ou similar de ferro/aço",
    },
    {
        "headings": {"8701", "8702", "8703", "8704", "8705", "8706"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por posição NCM — 87.01 a 87.06 (Lei 10.485/2002)",
        "descricao_padrao": "Trator/veículo automotor (posições 87.01 a 87.06)",
    },
    # --- Máquinas do Capítulo 84 — só autopropulsadas (§1º do art. 1º) ---
    {
        "headings": {"8429"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.29, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Máquina de terraplanagem/construção autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8432"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.32, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Máquina agrícola de preparação de solo, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8433"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.33, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Colheitadeira/máquina de colheita, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8434"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.34, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Ordenhadeira, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8435"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.35, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Prensa para vinho/sidra/sucos, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8436"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.36, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Outra máquina agrícola/florestal/avicultura, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "headings": {"8437"},
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por posição NCM — 84.37, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Máquina de limpeza/seleção de sementes/grãos, autopropulsada",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    # --- Autopeças — Anexo I da Lei 10.485/2002 (aplicação direta pelo NCM) ---
    {
        "headings": {"6813"},
        "categoria": "Autopeças",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo I)",
        "regra_identificacao": "Monofásico por posição NCM — 68.13 (Anexo I, Lei 10.485/2002)",
        "descricao_padrao": "Guarnições de fricção (lonas/pastilhas de freio)",
    },
    {
        "headings": {"8511"},
        "categoria": "Autopeças",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo I)",
        "regra_identificacao": "Monofásico por posição NCM — 85.11 (Anexo I, Lei 10.485/2002)",
        "descricao_padrao": "Aparelhos elétricos de ignição/arranque para motores",
    },
    {
        "headings": {"8707"},
        "categoria": "Autopeças",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo I)",
        "regra_identificacao": "Monofásico por posição NCM — 87.07 (Anexo I, Lei 10.485/2002)",
        "descricao_padrao": "Carroçarias para veículos automóveis",
    },
    {
        "headings": {"8708"},
        "categoria": "Autopeças",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo I)",
        "regra_identificacao": "Monofásico por posição NCM — 87.08 (Anexo I, Lei 10.485/2002)",
        "descricao_padrao": "Partes e acessórios de veículos automóveis (posições 87.01 a 87.05)",
    },
    # --- Autopeças — Anexo II da Lei 10.485/2002 (aplicação RESTRITA à
    # destinação específica — ver aviso). Lista completa dos 15 itens do
    # Anexo II, conforme texto oficial (planalto.gov.br/ccivil_03/leis/2002/l10485.htm).
    # Itens 1-2 e 9-10 são por posição/subposição; os demais NCM8 específicos
    # estão em _MONO_NCM_EXTRA e o item 14 (código fracionário) em
    # _MONO_PREFIX_RULES.
    {
        # Item 1 do Anexo II
        "headings": {"4009"},
        "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 1)",
        "regra_identificacao": "Monofásico por posição NCM — 40.09 (Anexo II, item 1, Lei 10.485/2002) — restrito",
        "descricao_padrao": "Tubos de borracha vulcanizada não endurecida, com acessórios",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 1 do Anexo II: tubos de borracha vulcanizada não "
            "endurecida da posição 40.09, com acessórios, próprios para "
            "máquinas e veículos autopropulsados das posições 84.29, "
            "8433.20, 8433.30.00, 8433.40.00, 8433.5, 87.01, 87.02, 87.03, "
            "87.04, 87.05 e 87.06.)"
        ),
    },
    {
        # Item 2 do Anexo II
        "headings": {"8431"},
        "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 2)",
        "regra_identificacao": "Monofásico por posição NCM — 84.31 (Anexo II, item 2, Lei 10.485/2002) — restrito",
        "descricao_padrao": "Partes da posição 84.31",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 2 do Anexo II: partes da posição 84.31, reconhecíveis "
            "como exclusiva ou principalmente destinadas às máquinas e "
            "aparelhos da posição 84.29.)"
        ),
    },
]

# Regras por subposição (6 dígitos) — usadas quando só uma fração da posição
# de 4 dígitos está sujeita à regra (ex.: 21.01 mistura café, que não entra
# no regime de bebidas frias, com chá/erva-mate, que entra).
_MONO_SUBHEADING_RULES = [
    {
        "subheading": "210120",
        "categoria": "Bebidas Frias",
        "referencia": "Lei 13.097/2015, art. 14",
        "regra_identificacao": "Monofásico por subposição NCM — 2101.20 (Lei 13.097/2015)",
        "descricao_padrao": "Extrato/concentrado de chá ou erva-mate para elaboração de bebida",
    },
    {
        "subheading": "842481",
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por subposição NCM — 8424.81, só autopropulsados (Lei 10.485/2002)",
        "descricao_padrao": "Aparelho mecânico de pulverização/dispersão, autopropulsado",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    {
        "subheading": "731029",
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por subposição NCM — 7310.29 (Lei 10.485/2002)",
        "descricao_padrao": "Outros pequenos reservatórios de ferro/aço",
    },
    # --- Autopeças — Anexo I (subposições de 6 dígitos) ---
    {"subheading": "840820", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8408.20 (Anexo I)", "descricao_padrao": "Motores diesel para propulsão de veículos"},
    {"subheading": "840991", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8409.91 (Anexo I)", "descricao_padrao": "Partes para motores de ignição por centelha"},
    {"subheading": "840999", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8409.99 (Anexo I)", "descricao_padrao": "Partes para motores diesel"},
    {"subheading": "841330", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8413.30 (Anexo I)", "descricao_padrao": "Bombas de combustível/óleo/refrigeração para motores"},
    {"subheading": "841520", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8415.20 (Anexo I)", "descricao_padrao": "Ar-condicionado para veículos"},
    {"subheading": "848310", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8483.10 (Anexo I)", "descricao_padrao": "Árvores de transmissão e manivelas"},
    {"subheading": "848330", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8483.30 (Anexo I)", "descricao_padrao": "Mancais e chumaceiras"},
    {"subheading": "848340", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8483.40 (Anexo I)", "descricao_padrao": "Engrenagens e rodas de fricção, caixas de marcha"},
    {"subheading": "848350", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8483.50 (Anexo I)", "descricao_padrao": "Volantes e polias"},
    {"subheading": "850520", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8505.20 (Anexo I)", "descricao_padrao": "Acoplamentos e embreagens eletromagnéticas"},
    {"subheading": "851220", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8512.20 (Anexo I)", "descricao_padrao": "Aparelhos de iluminação/sinalização visual para veículos"},
    {"subheading": "851240", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8512.40 (Anexo I)", "descricao_padrao": "Limpadores de para-brisa, degeladores e desembaçadores"},
    {"subheading": "852721", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8527.21 (Anexo I)", "descricao_padrao": "Autorrádios com sistema de descodificação"},
    {"subheading": "852729", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8527.29 (Anexo I)", "descricao_padrao": "Outros autorrádios"},
    {"subheading": "853910", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8539.10 (Anexo I)", "descricao_padrao": "Faróis e projetores em unidades seladas"},
]

# Regras por prefixo numérico arbitrário — para códigos que a própria lei
# cita abaixo do nível de 6 dígitos (ex.: "9032.89.2" no Anexo I, que cobre
# um grupo de NCM8 e não uma subposição padrão).
_MONO_PREFIX_RULES = [
    {
        "prefix": "9032892",
        "categoria": "Autopeças",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo I)",
        "regra_identificacao": "Monofásico — 9032.89.2 (Anexo I, Lei 10.485/2002)",
        "descricao_padrao": "Instrumento/aparelho automático de regulação ou controle para veículos",
    },
    {
        # Item 14 do Anexo II — "8483.60.1" é um código fracionário (abaixo
        # do nível de 6 dígitos) citado no texto oficial da lei.
        "prefix": "8483601",
        "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 14)",
        "regra_identificacao": "Monofásico — 8483.60.1 (Anexo II, item 14, Lei 10.485/2002) — restrito",
        "descricao_padrao": "Embreagens de fricção",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 14 do Anexo II: embreagens de fricção do código "
            "8483.60.1, próprias para máquinas dos códigos 84.29, 8433.20, "
            "8433.30.00, 8433.40.00 e 8433.5.)"
        ),
    },
]

_MONO_NCM_EXTRA = {
    "40141000": {
        "ncm": "40141000",
        "descricao": "Preservativos (artigos de borracha)",
        "categoria": "Produtos de Saúde",
        "referencia": "Lei 10.147/2000 art. 1º",
        "regra_identificacao": "Monofásico por NCM específico",
    },
    "21069010": {
        "ncm": "21069010",
        "descricao": "Preparação composta, não alcoólica, para elaboração de bebida (concentrado/xarope para refrigerante)",
        "categoria": "Bebidas Frias",
        "referencia": "Lei 13.097/2015, art. 14",
        "regra_identificacao": "Monofásico por NCM específico",
    },
    # --- Máquinas e veículos, Lei 10.485/2002 art. 1º ---
    "76129012": {
        "ncm": "76129012",
        "descricao": "Reservatório de alumínio",
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por NCM específico — 7612.90.12 (Lei 10.485/2002)",
    },
    "84306990": {
        "ncm": "84306990",
        "descricao": "Outras máquinas de terraplanagem/construção, autopropulsadas",
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º, §1º",
        "regra_identificacao": "Monofásico por NCM específico — 8430.69.90, só autopropulsados (Lei 10.485/2002)",
        "aviso": _AVISO_AUTOPROPULSADO,
    },
    "87162000": {
        "ncm": "87162000",
        "descricao": "Reboque agrícola autocarregável ou autodescarregável",
        "categoria": "Máquinas e Veículos",
        "referencia": "Lei 10.485/2002 art. 1º",
        "regra_identificacao": "Monofásico por NCM específico — 8716.20.00 (Lei 10.485/2002)",
    },
    # --- Autopeças — Anexo I (NCM8 específicos, aplicação direta) ---
    "40161010": {"ncm": "40161010", "descricao": "Batentes de borracha para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 4016.10.10 (Anexo I)"},
    "40169990": {
        "ncm": "40169990", "descricao": "Outras obras de borracha vulcanizada não endurecida (Anexo I, aplica-se apenas aos Ex 03 e Ex 05 da TIPI)",
        "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 4016.99.90, Ex 03 e 05 (Anexo I)",
        "aviso": "Este NCM só é monofásico nos Ex 03 e Ex 05 da TIPI — não no código inteiro. Confirme o Ex-tarifário exato do item antes de aplicar.",
    },
    "70071100": {"ncm": "70071100", "descricao": "Vidro de segurança temperado, para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 7007.11.00 (Anexo I)"},
    "70072100": {"ncm": "70072100", "descricao": "Vidro de segurança laminado, para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 7007.21.00 (Anexo I)"},
    "70091000": {"ncm": "70091000", "descricao": "Espelhos retrovisores para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 7009.10.00 (Anexo I)"},
    "73201000": {
        "ncm": "73201000", "descricao": "Molas de lâminas e suas folhas, de ferro/aço (Anexo I, aplica-se apenas ao Ex 01 da TIPI)",
        "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 7320.10.00, Ex 01 (Anexo I)",
        "aviso": "Este NCM só é monofásico no Ex 01 da TIPI — não no código inteiro. Confirme o Ex-tarifário exato do item antes de aplicar.",
    },
    "83012000": {"ncm": "83012000", "descricao": "Fechaduras para veículos automóveis", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8301.20.00 (Anexo I)"},
    "83023000": {"ncm": "83023000", "descricao": "Guarnições, ferragens e artigos semelhantes para veículos automóveis", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8302.30.00 (Anexo I)"},
    "84073390": {"ncm": "84073390", "descricao": "Outros motores de ignição por centelha, 250-1000cc, para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8407.33.90 (Anexo I)"},
    "84073490": {"ncm": "84073490", "descricao": "Outros motores de ignição por centelha, >1000cc, para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8407.34.90 (Anexo I)"},
    "84148021": {"ncm": "84148021", "descricao": "Turbocompressores para motores de veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8414.80.21 (Anexo I)"},
    "84148022": {"ncm": "84148022", "descricao": "Compressores de ar para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8414.80.22 (Anexo I)"},
    "84139100": {
        "ncm": "84139100", "descricao": "Partes de bombas para líquidos, para veículos (Anexo I, aplica-se apenas ao Ex 01 da TIPI)",
        "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8413.91.00, Ex 01 (Anexo I)",
        "aviso": "Este NCM só é monofásico no Ex 01 da TIPI — não no código inteiro. Confirme o Ex-tarifário exato do item antes de aplicar.",
    },
    "84213100": {"ncm": "84213100", "descricao": "Filtros de entrada de ar para motores de ignição por centelha/compressão", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8421.31.00 (Anexo I)"},
    "84314100": {"ncm": "84314100", "descricao": "Partes (baldes/caçambas/pás) de máquinas das posições 84.29/84.30", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8431.41.00 (Anexo I)"},
    "84314200": {"ncm": "84314200", "descricao": "Partes (lâminas/scarificadores) de máquinas das posições 84.29/84.30", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8431.42.00 (Anexo I)"},
    "84339090": {"ncm": "84339090", "descricao": "Outras partes de máquinas/aparelhos de colheita ou debulha (posição 84.33)", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8433.90.90 (Anexo I)"},
    "84818099": {
        "ncm": "84818099", "descricao": "Outras torneiras/válvulas e dispositivos semelhantes, para veículos (Anexo I, aplica-se apenas aos Ex 01 e Ex 02 da TIPI)",
        "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8481.80.99, Ex 01 e 02 (Anexo I)",
        "aviso": "Este NCM só é monofásico nos Ex 01 e Ex 02 da TIPI — não no código inteiro. Confirme o Ex-tarifário exato do item antes de aplicar.",
    },
    "84832000": {"ncm": "84832000", "descricao": "Rolamentos (mancais) de encaixe (plummer blocks) para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8483.20.00 (Anexo I)"},
    "85071000": {"ncm": "85071000", "descricao": "Acumuladores (baterias) de chumbo para arranque de motores", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8507.10.00 (Anexo I)"},
    "85123000": {"ncm": "85123000", "descricao": "Aparelhos de sinalização acústica (buzinas) para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8512.30.00 (Anexo I)"},
    "85129000": {"ncm": "85129000", "descricao": "Partes de aparelhos elétricos de iluminação/sinalização para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8512.90.00 (Anexo I)"},
    "85365090": {
        "ncm": "85365090", "descricao": "Outros interruptores para tensão até 1kV (Anexo I, aplica-se apenas ao Ex 01 da TIPI)",
        "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8536.50.90, Ex 01 (Anexo I)",
        "aviso": "Este NCM só é monofásico no Ex 01 da TIPI — não no código inteiro. Confirme o Ex-tarifário exato do item antes de aplicar.",
    },
    "85443000": {"ncm": "85443000", "descricao": "Jogos de fios para velas de ignição e outros jogos de fios para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 8544.30.00 (Anexo I)"},
    "90292010": {"ncm": "90292010", "descricao": "Taxímetros", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9029.20.10 (Anexo I)"},
    "90299010": {"ncm": "90299010", "descricao": "Partes de contadores/taxímetros/velocímetros para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9029.90.10 (Anexo I)"},
    "90303921": {"ncm": "90303921", "descricao": "Aparelhos de diagnóstico eletrônico para sistemas de veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9030.39.21 (Anexo I)"},
    "90318040": {"ncm": "90318040", "descricao": "Aparelhos e instrumentos de medida/controle para veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9031.80.40 (Anexo I)"},
    "91040000": {"ncm": "91040000", "descricao": "Relógios para painéis de instrumentos de veículos", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9104.00.00 (Anexo I)"},
    "94012000": {"ncm": "94012000", "descricao": "Assentos para veículos automóveis", "categoria": "Autopeças", "referencia": "Lei 10.485/2002, art. 3º (Anexo I)", "regra_identificacao": "Monofásico — 9401.20.00 (Anexo I)"},

    # --- Autopeças — Anexo II da Lei 10.485/2002 (itens 3-13 e 15; o item
    # 14 está em _MONO_PREFIX_RULES por ser um código fracionário) ---
    "84089090": {
        "ncm": "84089090", "descricao": "Motores diesel/semidiesel", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 3)", "regra_identificacao": "Monofásico — 8408.90.90 (Anexo II, item 3) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 3 do Anexo II: motores do código 8408.90.90, próprios "
            "para máquinas dos códigos 84.29, 8433.20, 8433.30.00, "
            "8433.40.00 e 8433.5.)"
        ),
    },
    "84122110": {
        "ncm": "84122110", "descricao": "Cilindros hidráulicos (motores hidráulicos de movimento retilíneo)", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 4)", "regra_identificacao": "Monofásico — 8412.21.10 (Anexo II, item 4) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 4 do Anexo II: cilindros hidráulicos do código "
            "8412.21.10, próprios para máquinas dos códigos 84.29, "
            "8433.20, 8433.30.00, 8433.40.00 e 8433.5.)"
        ),
    },
    "84122190": {
        "ncm": "84122190", "descricao": "Outros motores hidráulicos de movimento retilíneo (cilindros)", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 5)", "regra_identificacao": "Monofásico — 8412.21.90 (Anexo II, item 5) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 5 do Anexo II: outros motores hidráulicos de movimento "
            "retilíneo (cilindros) do código 8412.21.90, próprios para "
            "máquinas dos códigos 84.29, 8433.20, 8433.30.00, 8433.40.00 e "
            "8433.5.)"
        ),
    },
    "84123110": {
        "ncm": "84123110", "descricao": "Cilindros pneumáticos", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 6)", "regra_identificacao": "Monofásico — 8412.31.10 (Anexo II, item 6) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 6 do Anexo II: cilindros pneumáticos do código "
            "8412.31.10, próprios para produtos dos códigos 8701.20.00, "
            "87.02 e 87.04.)"
        ),
    },
    "84136019": {
        "ncm": "84136019", "descricao": "Bombas volumétricas rotativas", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 7)", "regra_identificacao": "Monofásico — 8413.60.19 (Anexo II, item 7) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 7 do Anexo II: bombas volumétricas rotativas do código "
            "8413.60.19, próprias para produtos dos códigos 84.29, "
            "8433.20, 8433.30.00, 8433.40.00, 8433.5, 8701.20.00, 87.02 e "
            "87.04.)"
        ),
    },
    "84148019": {
        "ncm": "84148019", "descricao": "Compressores de ar", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 8)", "regra_identificacao": "Monofásico — 8414.80.19 (Anexo II, item 8) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 8 do Anexo II: compressores de ar do código "
            "8414.80.19, próprios para produtos dos códigos 8701.20.00, "
            "87.02 e 87.04.)"
        ),
    },
    "84149039": {
        "ncm": "84149039", "descricao": "Caixas de ventilação para veículos autopropulsados", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 9)", "regra_identificacao": "Monofásico — 8414.90.39 (Anexo II, item 9) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 9 do Anexo II: caixas de ventilação para veículos "
            "autopropulsados, classificadas no código 8414.90.39.)"
        ),
    },
    "84329000": {
        "ncm": "84329000", "descricao": "Partes de máquinas de preparação/cultivo do solo", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 10)", "regra_identificacao": "Monofásico — 8432.90.00 (Anexo II, item 10) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 10 do Anexo II: partes classificadas no código "
            "8432.90.00, de máquinas das posições 8432.40.00 e "
            "8432.80.00.)"
        ),
    },
    "84811000": {
        "ncm": "84811000", "descricao": "Válvulas redutoras de pressão", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 11)", "regra_identificacao": "Monofásico — 8481.10.00 (Anexo II, item 11) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 11 do Anexo II: válvulas redutoras de pressão do "
            "código 8481.10.00, próprias para máquinas e veículos "
            "autopropulsados dos códigos 84.29, 8433.20, 8433.30.00, "
            "8433.40.00, 8433.5, 87.01, 87.02, 87.03, 87.04, 87.05 e "
            "87.06.)"
        ),
    },
    "84812090": {
        "ncm": "84812090", "descricao": "Válvulas para transmissões óleo-hidráulicas ou pneumáticas", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 12)", "regra_identificacao": "Monofásico — 8481.20.90 (Anexo II, item 12) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 12 do Anexo II: válvulas para transmissões "
            "óleo-hidráulicas ou pneumáticas do código 8481.20.90, "
            "próprias para máquinas dos códigos 84.29, 8433.20, "
            "8433.30.00, 8433.40.00 e 8433.5.)"
        ),
    },
    "84818092": {
        "ncm": "84818092", "descricao": "Válvulas solenoides", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 13)", "regra_identificacao": "Monofásico — 8481.80.92 (Anexo II, item 13) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 13 do Anexo II: válvulas solenoides do código "
            "8481.80.92, próprias para máquinas e veículos autopropulsados "
            "das posições 84.29, 8433.20, 8433.30.00, 8433.40.00, 8433.5, "
            "87.01, 87.02, 87.03, 87.04, 87.05 e 87.06.)"
        ),
    },
    "85011019": {
        "ncm": "85011019", "descricao": "Outros motores de corrente contínua, para acionamento elétrico de vidros de veículos autopropulsados", "categoria": "Autopeças (Anexo II)",
        "referencia": "Lei 10.485/2002, art. 3º (Anexo II, item 15)", "regra_identificacao": "Monofásico — 8501.10.19 (Anexo II, item 15) — restrito",
        "aviso": _AVISO_ANEXO_II + (
            " (Item 15 do Anexo II: outros motores de corrente contínua do "
            "código 8501.10.19, próprios para acionamento elétrico de "
            "vidros de veículos autopropulsados.)"
        ),
    },
}

# Índice rápido dos NCMs específicos da tabela principal
_MONOFASICOS_IDX = {m["ncm"]: m for m in MONOFASICOS}


def get_monofasico_info(ncm_clean):
    """
    Retorna (is_monofasico, mono_data) com campo 'regra_identificacao'.
    Prioridade: NCM específico > NCM extra > capítulo > posição > subposição.
    """
    # 1. NCM específico da tabela principal
    if ncm_clean in _MONOFASICOS_IDX:
        data = dict(_MONOFASICOS_IDX[ncm_clean])
        data["regra_identificacao"] = "Monofásico por NCM específico"
        data.setdefault("ncm_formatado", format_ncm(ncm_clean))
        data.setdefault("aviso", None)
        return True, data

    # 2. NCMs extras (ex.: preservativos, autopeças do Anexo I com "Ex" específico)
    if ncm_clean in _MONO_NCM_EXTRA:
        data = dict(_MONO_NCM_EXTRA[ncm_clean])
        data.setdefault("ncm_formatado", format_ncm(ncm_clean))
        data.setdefault("aviso", None)
        return True, data

    chapter = ncm_clean[:2]
    heading = ncm_clean[:4]
    subheading = ncm_clean[:6]

    # 3. Regras por capítulo
    for rule in _MONO_CHAPTER_RULES:
        if rule["chapter"] != chapter:
            continue
        if rule["headings"] is not None and heading not in rule["headings"]:
            continue
        if ncm_clean in rule.get("excecoes", ()):
            continue
        return True, {
            "ncm": ncm_clean,
            "ncm_formatado": format_ncm(ncm_clean),
            "descricao": rule["descricao_padrao"],
            "categoria": rule["categoria"],
            "referencia": rule["referencia"],
            "regra_identificacao": rule["regra_identificacao"],
            "aviso": rule.get("aviso"),
        }

    # 4. Regras por posição (heading)
    for rule in _MONO_HEADING_RULES:
        if heading not in rule["headings"]:
            continue
        return True, {
            "ncm": ncm_clean,
            "ncm_formatado": format_ncm(ncm_clean),
            "descricao": rule["descricao_padrao"],
            "categoria": rule["categoria"],
            "referencia": rule["referencia"],
            "regra_identificacao": rule["regra_identificacao"],
            "aviso": rule.get("aviso"),
        }

    # 5. Regras por subposição (6 dígitos)
    for rule in _MONO_SUBHEADING_RULES:
        if subheading != rule["subheading"]:
            continue
        return True, {
            "ncm": ncm_clean,
            "ncm_formatado": format_ncm(ncm_clean),
            "descricao": rule["descricao_padrao"],
            "categoria": rule["categoria"],
            "referencia": rule["referencia"],
            "regra_identificacao": rule["regra_identificacao"],
            "aviso": rule.get("aviso"),
        }

    # 6. Regras por prefixo numérico arbitrário (para códigos "quebrados" que a
    # lei cita abaixo do nível de 6 dígitos, ex.: 9032.89.2)
    for rule in _MONO_PREFIX_RULES:
        if not ncm_clean.startswith(rule["prefix"]):
            continue
        return True, {
            "ncm": ncm_clean,
            "ncm_formatado": format_ncm(ncm_clean),
            "descricao": rule["descricao_padrao"],
            "categoria": rule["categoria"],
            "referencia": rule["referencia"],
            "regra_identificacao": rule["regra_identificacao"],
            "aviso": rule.get("aviso"),
        }

    return False, None


# --- Tabela de alíquotas IPI por capítulo (simplificado) ---
IPI_RATES = {
    "01": 0.0, "02": 0.0, "03": 0.0, "04": 0.0, "05": 0.0,
    "06": 0.0, "07": 0.0, "08": 0.0, "09": 0.0, "10": 0.0,
    "11": 0.0, "12": 0.0, "13": 0.0, "14": 0.0, "15": 0.0,
    "16": 0.0, "17": 0.0, "18": 5.0, "19": 0.0, "20": 0.0,
    "21": 0.0, "22": 15.0, "23": 0.0, "24": 15.0, "25": 0.0,
    "26": 0.0, "27": 0.0, "28": 5.0, "29": 5.0, "30": 0.0,
    "31": 0.0, "32": 5.0, "33": 15.0, "34": 5.0, "35": 5.0,
    "36": 5.0, "37": 5.0, "38": 5.0, "39": 5.0, "40": 3.0,
    "41": 5.0, "42": 10.0, "43": 5.0, "44": 5.0, "45": 5.0,
    "46": 5.0, "47": 0.0, "48": 5.0, "49": 0.0, "50": 10.0,
    "51": 10.0, "52": 10.0, "53": 10.0, "54": 10.0, "55": 10.0,
    "56": 10.0, "57": 5.0, "58": 10.0, "59": 10.0, "60": 10.0,
    "61": 15.0, "62": 15.0, "63": 10.0, "64": 15.0, "65": 5.0,
    "66": 5.0, "67": 5.0, "68": 5.0, "69": 5.0, "70": 5.0,
    "71": 5.0, "72": 0.0, "73": 5.0, "74": 5.0, "75": 5.0,
    "76": 5.0, "77": 0.0, "78": 5.0, "79": 5.0, "80": 5.0,
    "81": 5.0, "82": 5.0, "83": 5.0, "84": 5.0, "85": 10.0,
    "86": 5.0, "87": 25.0, "88": 5.0, "89": 0.0, "90": 5.0,
    "91": 15.0, "92": 15.0, "93": 10.0, "94": 5.0, "95": 10.0,
    "96": 5.0, "97": 0.0, "98": 0.0, "99": 0.0,
}

# --- ClassTrib por NCM (exemplos) ---
CLASS_TRIB = {
    "30": {"class": "Medicamentos", "pis_cst": "04", "cofins_cst": "04", "ipi_cst": "52"},
    "33": {"class": "Cosméticos/Perfumes", "pis_cst": "04", "cofins_cst": "04", "ipi_cst": "50"},
    "22": {"class": "Bebidas", "pis_cst": "03", "cofins_cst": "03", "ipi_cst": "50"},
    "27": {"class": "Combustíveis", "pis_cst": "04", "cofins_cst": "04", "ipi_cst": "53"},
    "87": {"class": "Veículos Automotores", "pis_cst": "04", "cofins_cst": "04", "ipi_cst": "50"},
    "84": {"class": "Máquinas e Aparelhos", "pis_cst": "01", "cofins_cst": "01", "ipi_cst": "50"},
    "85": {"class": "Materiais Elétricos", "pis_cst": "01", "cofins_cst": "01", "ipi_cst": "50"},
}

# --- ICMS por estado (alíquota interna) ---
ICMS_STATES = [
    {"uf": "AC", "estado": "Acre", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "AL", "estado": "Alagoas", "aliq_interna": 19.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "AP", "estado": "Amapá", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "AM", "estado": "Amazonas", "aliq_interna": 20.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": "ZFM - regime especial"},
    {"uf": "BA", "estado": "Bahia", "aliq_interna": 19.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "CE", "estado": "Ceará", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "DF", "estado": "Distrito Federal", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "ES", "estado": "Espírito Santo", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "GO", "estado": "Goiás", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "MA", "estado": "Maranhão", "aliq_interna": 20.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "MT", "estado": "Mato Grosso", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "MS", "estado": "Mato Grosso do Sul", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "MG", "estado": "Minas Gerais", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "PA", "estado": "Pará", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "PB", "estado": "Paraíba", "aliq_interna": 20.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "PR", "estado": "Paraná", "aliq_interna": 19.5, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "PE", "estado": "Pernambuco", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "PI", "estado": "Piauí", "aliq_interna": 22.5, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": "Lei nº 8.558/2024 · vigente desde 01/04/2025"},
    {"uf": "RJ", "estado": "Rio de Janeiro", "aliq_interna": 20.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "RN", "estado": "Rio Grande do Norte", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "RS", "estado": "Rio Grande do Sul", "aliq_interna": 17.5, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "RO", "estado": "Rondônia", "aliq_interna": 17.5, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "RR", "estado": "Roraima", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "SC", "estado": "Santa Catarina", "aliq_interna": 17.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "SP", "estado": "São Paulo", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 12.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "SE", "estado": "Sergipe", "aliq_interna": 19.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
    {"uf": "TO", "estado": "Tocantins", "aliq_interna": 18.0, "aliq_interestadual_sul_sudeste": 7.0, "aliq_interestadual_outros": 12.0, "aliq_importado": 4.0, "reducao_bc": "Varia por produto", "obs": ""},
]

def strip_html(text):
    return re.sub(r'<[^>]+>', '', text) if text else text

def format_ncm(ncm):
    ncm = re.sub(r'\D', '', ncm)
    if len(ncm) == 8:
        return f"{ncm[:4]}.{ncm[4:6]}.{ncm[6:]}"
    return ncm

def get_ipi_rate(ncm_code):
    chapter = ncm_code[:2]
    return IPI_RATES.get(chapter, 0.0)

def get_class_trib(ncm_code):
    chapter = ncm_code[:2]
    return CLASS_TRIB.get(chapter, {
        "class": "Produto em Geral",
        "pis_cst": "01",
        "cofins_cst": "01",
        "ipi_cst": "50"
    })


# --- Alíquota zero de PIS/COFINS — motor de regras ---
#
# Dois módulos distintos:
#   - "Hortifrúti"/"Papel Jornal": art. 28 da Lei nº 10.865/2004 (nota de
#     exibição fixa, mantida como já estava antes deste módulo).
#   - "Cesta Básica": art. 1º da Lei nº 10.925/2004, incisos I a XXVII
#     (a lei não tem incisos VIII e XVII). Nota padronizada:
#     "Alíquota Zero - Cesta Básica (Lei 10.925/2004, art. 1º, inciso X)".
#
# Regras mais específicas (ncm8/subposição/prefixo) vêm ANTES das mais
# genéricas (capítulo) na lista, para que a primeira regra que bater
# "vença" — importante porque o feijão (Cap. 7, inciso V da Cesta Básica)
# precisa ser resolvido antes da regra genérica de hortifrúti (Cap. 7/8).
_ALIQ_ZERO_RULES = [
    # ---- Cesta Básica — itens específicos que colidiriam com regras mais
    # amplas (hortifrúti) se checados depois ----
    {
        "level": "ncm8", "codes": {"07133319", "07133329", "07133399"},
        "modulo": "Cesta Básica", "inciso": "V", "descricao": "feijão",
    },

    # ---- Hortifrúti e Papel Jornal — Lei nº 10.865/2004 (inalterado) ----
    {
        "level": "chapter", "codes": {"07", "08"},
        "modulo": "Hortifrúti",
        "nota_fixa": (
            'Alíquota zero conforme art. 28, III da Lei 10.865/2004 - '
            'produtos hortícolas e frutas in natura'
        ),
        "base_legal_fixa": "Art. 28, III da Lei nº 10.865/2004",
    },
    {
        "level": "heading", "codes": {"0407"},
        "modulo": "Hortifrúti",
        "nota_fixa": (
            'Alíquota zero conforme art. 28, III da Lei 10.865/2004 - '
            'produtos hortícolas e frutas in natura'
        ),
        "base_legal_fixa": "Art. 28, III da Lei nº 10.865/2004",
    },
    {
        "level": "heading", "codes": {"4801"},
        "modulo": "Papel Jornal",
        "nota_fixa": (
            'Alíquota zero conforme art. 28, I da Lei 10.865/2004 - '
            'papel destinado à impressão de jornais. Benefício foi '
            'instituído por prazo determinado e prorrogado por '
            'legislação posterior — confirme a vigência atual antes '
            'de aplicar'
        ),
        "base_legal_fixa": "Art. 28, I da Lei nº 10.865/2004",
    },

    # ---- Cesta Básica — Lei nº 10.925/2004, art. 1º ----
    # Incisos I-III já existiam como módulo separado ("insumos
    # agropecuários"); mantidos com a mesma regra de NCM, só relabelados
    # para o padrão único "Cesta Básica" pedido.
    {
        "level": "chapter", "codes": {"31"},
        "modulo": "Cesta Básica", "inciso": "I", "descricao": "adubos/fertilizantes (Capítulo 31)",
        "aviso_extra": "Exceto produtos de uso veterinário — confirme se este NCM específico se enquadra na exceção antes de aplicar.",
    },
    {
        "level": "heading", "codes": {"3808"},
        "modulo": "Cesta Básica", "inciso": "II", "descricao": "defensivos agropecuários (posição 38.08)",
    },
    {
        "level": "heading", "codes": {"1209"},
        "modulo": "Cesta Básica", "inciso": "III", "descricao": "sementes para semeadura (posição 12.09)",
    },
    {
        "level": "heading", "codes": {"0602"},
        "modulo": "Cesta Básica", "inciso": "III", "descricao": "mudas para plantio (posição 06.02)",
    },
    {
        "level": "chapter", "codes": {"25"},
        "modulo": "Cesta Básica", "inciso": "IV", "descricao": "corretivo de solo de origem mineral (Capítulo 25)",
    },
    {
        "level": "subheading", "codes": {"100620"},
        "modulo": "Cesta Básica", "inciso": "V", "descricao": "arroz descascado (integral)",
    },
    {
        "level": "subheading", "codes": {"100630"},
        "modulo": "Cesta Básica", "inciso": "V", "descricao": "arroz semibranqueado ou branqueado",
    },
    {
        "level": "subheading", "codes": {"110620"},
        "modulo": "Cesta Básica", "inciso": "V", "descricao": "farinha (posição 11.06)",
    },
    {
        "level": "ncm8", "codes": {"30029099"},
        "modulo": "Cesta Básica", "inciso": "VI", "descricao": "inoculantes agrícolas",
    },
    {
        "level": "subheading", "codes": {"300230"},
        "modulo": "Cesta Básica", "inciso": "VII", "descricao": "vacinas veterinárias (NCM 3002.30)",
    },
    {
        "level": "subheading", "codes": {"110220", "110313", "110419"},
        "modulo": "Cesta Básica", "inciso": "IX", "descricao": "farinha/grumos/sêmolas de milho",
    },
    {
        "level": "subheading", "codes": {"010511"},
        "modulo": "Cesta Básica", "inciso": "X", "descricao": "pintos de 1 dia",
    },
    # XI - leite (fluido, em pó exceto condensado, fermentado, bebidas e
    # compostos lácteos, fórmulas infantis)
    {
        "level": "heading", "codes": {"0401"},
        "modulo": "Cesta Básica", "inciso": "XI", "descricao": "leite fluido pasteurizado/ultrapasteurizado",
    },
    {
        "level": "subheading", "codes": {"040210", "040221", "040229"},
        "modulo": "Cesta Básica", "inciso": "XI", "descricao": "leite em pó (integral, semidesnatado ou desnatado)",
        "aviso_extra": "Leite condensado (NCM 0402.91/0402.99) NÃO tem esse benefício — já foi objeto de solução de consulta da Receita Federal negando a alíquota zero.",
    },
    {
        "level": "heading", "codes": {"0403"},
        "modulo": "Cesta Básica", "inciso": "XI", "descricao": "leite fermentado",
    },
    {
        "level": "subheading", "codes": {"040490"},
        "modulo": "Cesta Básica", "inciso": "XI", "descricao": "bebidas e compostos lácteos",
    },
    {
        "level": "subheading", "codes": {"190110"},
        "modulo": "Cesta Básica", "inciso": "XI", "descricao": "fórmulas infantis",
    },
    {
        "level": "subheading", "codes": {"040610"},
        "modulo": "Cesta Básica", "inciso": "XII", "descricao": "queijo fresco/requeijão",
        "aviso_extra": (
            "Único código do Capítulo 04 com correspondência direta e "
            "confirmada por solução de consulta COSIT para o inciso XII "
            "(demais tipos citados na lei — mussarela, minas, prato, "
            "coalho, ricota, provolone, parmesão, do reino — não têm "
            "código de 8 dígitos exclusivo; ver aviso da posição 04.06)."
        ),
    },
    {
        "level": "subheading", "codes": {"040410"},
        "modulo": "Cesta Básica", "inciso": "XIII", "descricao": "soro de leite fluido (para industrialização de consumo humano)",
    },
    {
        "level": "ncm8", "codes": {"11010010"},
        "modulo": "Cesta Básica", "inciso": "XIV", "descricao": "farinha de trigo",
    },
    {
        "level": "heading", "codes": {"1001"},
        "modulo": "Cesta Básica", "inciso": "XV", "descricao": "trigo",
    },
    {
        "level": "ncm8", "codes": {"19012000", "19059090"},
        "modulo": "Cesta Básica", "inciso": "XVI", "descricao": "pré-mistura para pão comum e pão comum",
        "aviso_extra": (
            "Aplica-se apenas ao Ex 01 da TIPI de cada código — não ao "
            "código inteiro. Sujeito a interpretação restritiva conforme "
            "jurisprudência do STJ (pão comum = feito exclusivamente com "
            "farinha de trigo) - confirme composição do produto."
        ),
    },
    {
        "level": "heading", "codes": {"1902"},
        "modulo": "Cesta Básica", "inciso": "XVIII", "descricao": "massas alimentícias",
        "aviso_extra": "Sujeito a interpretação restritiva conforme jurisprudência do STJ - confirme composição do produto.",
    },
    # XIX - carnes
    {
        "level": "heading", "codes": {"0201", "0202"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "carne bovina fresca/refrigerada/congelada",
    },
    {
        "level": "ncm8", "codes": {"02061000"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "miudezas bovinas frescas/refrigeradas",
    },
    {
        "level": "prefix", "codes": {"02062"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "miudezas bovinas congeladas (0206.2)",
    },
    {
        "level": "ncm8", "codes": {"02102000"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "carne bovina salgada/em salmoura/seca/defumada",
    },
    {
        "level": "ncm8", "codes": {"05069000"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "ossos e núcleos córneos não trabalhados",
    },
    {
        "level": "ncm8", "codes": {"05100010"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "glândulas e outros produtos de origem animal",
    },
    {
        "level": "prefix", "codes": {"1502101"},
        "modulo": "Cesta Básica", "inciso": "XIX-a", "descricao": "gorduras de bovinos/ovinos/caprinos, em bruto (1502.10.1)",
    },
    {
        "level": "heading", "codes": {"0203"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "carne suína fresca/refrigerada/congelada",
    },
    {
        "level": "ncm8", "codes": {"02063000"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "miudezas suínas frescas/refrigeradas",
    },
    {
        "level": "prefix", "codes": {"02064"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "miudezas suínas congeladas (0206.4)",
    },
    {
        "level": "heading", "codes": {"0207", "0209"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "carnes/miudezas de aves; toucinho e gorduras de porco/aves",
    },
    {
        "level": "prefix", "codes": {"02101"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "carne suína salgada/em salmoura/seca/defumada (0210.1)",
    },
    {
        "level": "ncm8", "codes": {"02109900"},
        "modulo": "Cesta Básica", "inciso": "XIX-b", "descricao": "carne de frango salgada/em salmoura/seca/defumada",
    },
    {
        "level": "heading", "codes": {"0204"},
        "modulo": "Cesta Básica", "inciso": "XIX-c", "descricao": "carne de ovinos/caprinos fresca/refrigerada/congelada",
    },
    {
        "level": "ncm8", "codes": {"02068000"},
        "modulo": "Cesta Básica", "inciso": "XIX-c", "descricao": "miudezas de ovinos/caprinos",
    },
    # XX - peixes
    {
        "level": "heading", "codes": {"0302"}, "excecoes": {"03029000"},
        "modulo": "Cesta Básica", "inciso": "XX-a", "descricao": "peixes frescos ou refrigerados (exceto código 0302.90.00)",
    },
    {
        "level": "heading", "codes": {"0303", "0304"},
        "modulo": "Cesta Básica", "inciso": "XX-b", "descricao": "peixes congelados; filés e outras carnes de peixes",
    },
    {
        "level": "heading", "codes": {"0901"},
        "modulo": "Cesta Básica", "inciso": "XXI", "descricao": "café",
    },
    {
        "level": "prefix", "codes": {"21011"},
        "modulo": "Cesta Básica", "inciso": "XXI", "descricao": "extratos/essências/concentrados de café (2101.1)",
    },
    {
        "level": "ncm8", "codes": {"17011400", "17019900"},
        "modulo": "Cesta Básica", "inciso": "XXII", "descricao": "açúcar",
    },
    {
        "level": "heading", "codes": {"1507", "1508", "1509", "1510", "1511", "1512", "1513", "1514"},
        "modulo": "Cesta Básica", "inciso": "XXIII", "descricao": "óleo de soja e outros óleos vegetais",
    },
    {
        "level": "ncm8", "codes": {"04051000"},
        "modulo": "Cesta Básica", "inciso": "XXIV", "descricao": "manteiga",
    },
    {
        "level": "ncm8", "codes": {"15171000"},
        "modulo": "Cesta Básica", "inciso": "XXV", "descricao": "margarina",
    },
    {
        "level": "ncm8", "codes": {"34011190"},
        "modulo": "Cesta Básica", "inciso": "XXVI", "descricao": "sabões de toucador",
        "aviso_extra": (
            "Aplica-se apenas ao Ex 01 da TIPI (sabão de toucador) — não "
            "ao código inteiro. Este NCM também consta na tabela de "
            "produtos monofásicos (Lei 10.147/2000) para outros Ex da "
            "TIPI não abrangidos por este benefício; confirme qual regime "
            "se aplica ao Ex específico do seu produto."
        ),
    },
    {
        "level": "heading", "codes": {"3306"},
        "modulo": "Cesta Básica", "inciso": "XXVII", "descricao": "produtos de higiene bucal/dentária",
    },
]


def get_pis_cofins_aliquota_zero(ncm_clean):
    """
    Alíquota zero de PIS/COFINS.

    Dois módulos: "Hortifrúti"/"Papel Jornal" (art. 28 da Lei 10.865/2004)
    e "Cesta Básica" (art. 1º da Lei 10.925/2004, incisos I a XXVII — a lei
    não tem incisos VIII e XVII). Ver _ALIQ_ZERO_RULES para o detalhamento
    por inciso/NCM.
    """
    chapter = ncm_clean[:2]
    heading = ncm_clean[:4]
    subheading = ncm_clean[:6]

    for rule in _ALIQ_ZERO_RULES:
        level = rule["level"]
        if level == "chapter" and chapter not in rule["codes"]:
            continue
        if level == "heading" and heading not in rule["codes"]:
            continue
        if level == "subheading" and subheading not in rule["codes"]:
            continue
        if level == "ncm8" and ncm_clean not in rule["codes"]:
            continue
        if level == "prefix" and not any(ncm_clean.startswith(p) for p in rule["codes"]):
            continue
        if ncm_clean in rule.get("excecoes", ()):
            continue

        if "nota_fixa" in rule:
            return {
                'aliquota_zero': True,
                'base_legal': rule['base_legal_fixa'],
                'nota': rule['nota_fixa'],
                'modulo': rule['modulo'],
                'inciso': rule.get('inciso'),
            }

        base_legal = f"Lei 10.925/2004, art. 1º, inciso {rule['inciso']}"
        nota = f"Alíquota Zero - Cesta Básica ({base_legal}) - {rule['descricao']}"
        if rule.get("aviso_extra"):
            nota += f". {rule['aviso_extra']}"
        return {
            'aliquota_zero': True,
            'base_legal': base_legal,
            'nota': nota,
            'modulo': rule['modulo'],
            'inciso': rule['inciso'],
        }

    return None


def get_icms_hortifruti_aviso(ncm_clean):
    """
    Aviso sobre isenção/redução de ICMS para hortifrutigranjeiros in natura.
    Não é regra federal uniforme — varia por estado, então apenas alerta
    o usuário em vez de zerar automaticamente a alíquota.
    """
    chapter = ncm_clean[:2]
    if chapter in ('07', '08'):
        return (
            'Produtos hortifrutigranjeiros in natura frequentemente possuem '
            'isenção ou redução de base de cálculo de ICMS, conforme legislação '
            'de cada estado (geralmente com base no Convênio ICMS 44/75 e suas '
            'alterações). Consulte a legislação estadual específica antes de '
            'aplicar a alíquota padrão.'
        )
    return None


def get_queijo_aviso(ncm_clean, descricao):
    """
    Avisos sobre queijos (posição 04.06) que NÃO se enquadram
    automaticamente na alíquota zero do inciso XII da Lei 10.925/2004
    (0406.10 já é tratado como alíquota zero em _ALIQ_ZERO_RULES).
    """
    heading = ncm_clean[:4]
    subheading = ncm_clean[:6]
    descricao_lower = (descricao or '').lower()

    if heading != '0406':
        return None

    if subheading == '040620':
        return (
            'Queijo ralado ou em pó (NCM 0406.20) está EXCLUÍDO do '
            'benefício do inciso XII, mesmo sendo queijo — a Receita '
            'Federal já negou expressamente esse benefício para queijos '
            'ralados/em pó (ex.: tipo tropical/regional do norte).'
        )

    if subheading == '040630':
        return (
            'Queijo fundido (NCM 0406.30) está fora do escopo do inciso '
            'XII da Lei 10.925/2004 — não tem alíquota zero.'
        )

    if 'cheddar' in descricao_lower:
        return (
            'Queijo tipo cheddar NÃO tem alíquota zero — a Receita '
            'Federal já negou expressamente esse benefício para esse tipo '
            'de queijo, mesmo quando classificado na posição 04.06.'
        )

    if 'cream cheese' in descricao_lower or 'queijo cremoso' in descricao_lower:
        return (
            'Cream cheese (queijo cremoso) NÃO tem alíquota zero — a '
            'Receita Federal já negou expressamente esse benefício para '
            'esse tipo de queijo, mesmo quando classificado na posição '
            '04.06.'
        )

    if subheading != '040610':
        return (
            'Este NCM (posição 04.06) pode abranger tipos de queijo '
            'citados no inciso XII (mussarela, minas, prato, coalho, '
            'ricota, provolone, parmesão, do reino), mas a Receita Federal '
            '(solução de consulta COSIT) esclarece que a lei não define um '
            'código de 8 dígitos exclusivo por tipo de queijo — a '
            'aplicação da alíquota zero depende da composição/classificação '
            'técnica do produto, analisada caso a caso. Este NCM NÃO foi '
            'marcado como alíquota zero automaticamente; confirme o '
            'enquadramento antes de aplicar o benefício.'
        )

    return None


def get_icms_piaui(ncm_clean):
    """
    Alíquota ICMS interna e ST do Piauí conforme Lei nº 8.558/2024
    vigente desde 01/04/2025.
    """
    chapter = ncm_clean[:2]

    # 27% — bebidas alcoólicas, cigarros, armas, fogos de artifício
    if chapter == '22':
        aliquota, categoria = 27.0, 'Bebidas alcoólicas'
        tem_st = True
        obs_st = 'ST prevista — Protocolo ICMS 11/1991 e Convênio ICMS 52/2017'
    elif chapter == '24':
        aliquota, categoria = 27.0, 'Cigarros e derivados do tabaco'
        tem_st = True
        obs_st = 'ST prevista — Convênio ICMS 37/1994'
    elif chapter == '93':
        aliquota, categoria = 27.0, 'Armas e munições'
        tem_st = False
        obs_st = None
    elif chapter == '36':
        aliquota, categoria = 27.0, 'Fogos de artifício e explosivos'
        tem_st = False
        obs_st = None

    # 25% — combustíveis, energia elétrica (>200kWh), telecomunicações
    elif chapter == '27':
        aliquota, categoria = 25.0, 'Combustíveis e lubrificantes'
        tem_st = True
        obs_st = 'ST obrigatória — Convênio ICMS 110/2007'

    # 12% — cesta básica, insumos agropecuários, medicamentos essenciais
    elif chapter in ['01','02','03','04','05','07','08','09','10','11',
                     '12','13','14','15','16','17','19','20','21']:
        aliquota, categoria = 12.0, 'Alimentos e produtos da cesta básica'
        tem_st = False
        obs_st = None
    elif chapter == '30':
        aliquota, categoria = 12.0, 'Medicamentos essenciais'
        tem_st = True
        obs_st = 'ST prevista — Protocolo ICMS 26/1985 (verificar lista de isentos)'
    elif chapter == '31':
        aliquota, categoria = 12.0, 'Insumos agropecuários'
        tem_st = False
        obs_st = None

    # Isentos — livros, jornais, periódicos
    elif chapter == '49':
        aliquota, categoria = 0.0, 'Livros, jornais e periódicos (isenção)'
        tem_st = False
        obs_st = None

    # Regra geral com ST conhecida
    elif chapter == '33':
        aliquota, categoria = 22.5, 'Perfumaria e cosméticos'
        tem_st = True
        obs_st = 'ST prevista — Protocolo ICMS 190/2009'
    elif chapter == '40':
        aliquota, categoria = 22.5, 'Pneus e borracha'
        tem_st = True
        obs_st = 'ST prevista — Convênio ICMS 85/1993'
    elif chapter == '85':
        aliquota, categoria = 22.5, 'Materiais elétricos e eletrônicos'
        tem_st = True
        obs_st = 'ST prevista — Protocolo ICMS 27/2009 (verificar itens)'
    elif chapter == '87':
        aliquota, categoria = 22.5, 'Veículos automotores'
        tem_st = True
        obs_st = 'ST prevista — Convênio ICMS 132/1992'

    # Regra geral: 22,5%
    else:
        aliquota, categoria = 22.5, 'Produto em geral — alíquota padrão'
        tem_st = False
        obs_st = 'Verificar Convênios ICMS e Protocolos CONFAZ do Piauí'

    return {
        'aliquota': aliquota,
        'categoria': categoria,
        'tem_st': tem_st,
        'obs_st': obs_st,
        'base_legal': 'Lei Estadual nº 8.558/2024',
        'vigencia': '01/04/2025',
        'chapter': chapter,
        'faixas': [
            {'aliquota': 27.0, 'descricao': 'Bebidas alcoólicas, cigarros, armas, fogos de artifício', 'caps': '22, 24, 36, 93'},
            {'aliquota': 25.0, 'descricao': 'Energia elétrica (>200 kWh), combustíveis, telecomunicações', 'caps': '27'},
            {'aliquota': 22.5, 'descricao': 'Demais mercadorias — alíquota geral', 'caps': 'demais capítulos'},
            {'aliquota': 12.0, 'descricao': 'Cesta básica, insumos agropecuários, medicamentos essenciais', 'caps': '01-21, 30, 31'},
            {'aliquota': 0.0,  'descricao': 'Isentos: livros, jornais, periódicos, exportações', 'caps': '49'},
        ],
    }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/ncm/<ncm_code>')
def get_ncm(ncm_code):
    ncm_clean = re.sub(r'\D', '', ncm_code)
    if len(ncm_clean) not in [8]:
        ncm_clean = ncm_clean.zfill(8)

    try:
        resp = requests.get(
            f"https://brasilapi.com.br/api/ncm/v1/{ncm_clean}",
            timeout=10,
            headers={"User-Agent": "SistemaNcm/1.0"}
        )
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list):
            ncm_data = data[0] if data else {}
        else:
            ncm_data = data
    except requests.exceptions.HTTPError as e:
        if e.response and e.response.status_code == 404:
            return jsonify({"error": "NCM não encontrado na base de dados"}), 404
        return jsonify({"error": f"Erro ao consultar BrasilAPI: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": f"Erro de conexão: {str(e)}"}), 500

    ipi_rate = get_ipi_rate(ncm_clean)
    class_trib = get_class_trib(ncm_clean)
    chapter = ncm_clean[:2]

    is_monofasico, mono_data = get_monofasico_info(ncm_clean)
    pis_cofins_zero = get_pis_cofins_aliquota_zero(ncm_clean)

    # Quando a alíquota zero se aplica, ela prevalece na exibição sobre a
    # classificação monofásica — podem coexistir em teoria (ex.: dois "Ex"
    # distintos do mesmo NCM8 com regimes diferentes), mas mostrar as duas
    # etiquetas ao mesmo tempo confundiria o usuário. A reforma tributária
    # (LC 214/2025) é um regime independente e usa is_monofasico original.
    is_monofasico_display = False if pis_cofins_zero else is_monofasico

    if pis_cofins_zero:
        pis_cst = cofins_cst = "06"
    elif is_monofasico:
        # CST "03" (alíquota por unidade de medida) é específico do regime de
        # bebidas frias (Lei 13.097/2015) — inclui o capítulo 22 inteiro (por
        # ora mantém vinhos/destilados no mesmo CST histórico, já que seu
        # enquadramento não foi reauditado) e concentrados/xaropes do
        # capítulo 21 classificados como "Bebidas Frias". Os demais setores
        # monofásicos (medicamentos, perfumaria, combustíveis, veículos,
        # autopeças, pneus) usam CST "04" (revenda a alíquota zero).
        pis_cst = cofins_cst = (
            "03" if chapter == "22" or mono_data.get("categoria") == "Bebidas Frias"
            else "04"
        )
    else:
        pis_cst = cofins_cst = "01"
    pis_aliq_lr = 0.0 if pis_cofins_zero else 1.65
    pis_aliq_lp = 0.0 if pis_cofins_zero else 0.65
    cofins_aliq_lr = 0.0 if pis_cofins_zero else 7.6
    cofins_aliq_lp = 0.0 if pis_cofins_zero else 3.0
    nota_aliquota_zero = pis_cofins_zero["nota"] if pis_cofins_zero else None
    descricao_stripped = strip_html(ncm_data.get("descricao", "Não encontrado"))

    result = {
        "ncm": ncm_clean,
        "ncm_formatado": format_ncm(ncm_clean),
        "descricao": descricao_stripped,
        "data_inicio": ncm_data.get("data_inicio", ""),
        "data_fim": ncm_data.get("data_fim", ""),
        "tipo_ato": ncm_data.get("tipo_ato", ""),
        "numero_ato": ncm_data.get("numero_ato", ""),
        "ano_ato": ncm_data.get("ano_ato", ""),
        "tributacao": {
            "ipi": {
                "aliquota": ipi_rate,
                "base_calculo": "Valor da operação + frete + seguro",
                "incidencia": "Saída do estabelecimento industrial/equiparado",
                "regime": "TIPI - Tabela de Incidência do IPI",
            },
            "pis": {
                "lucro_real": {"aliquota": pis_aliq_lr, "cst": pis_cst},
                "lucro_presumido": {"aliquota": pis_aliq_lp, "cst": pis_cst},
                "monofasico": is_monofasico_display,
                "aliquota_zero": bool(pis_cofins_zero),
                "nota_aliquota_zero": nota_aliquota_zero,
            },
            "cofins": {
                "lucro_real": {"aliquota": cofins_aliq_lr, "cst": cofins_cst},
                "lucro_presumido": {"aliquota": cofins_aliq_lp, "cst": cofins_cst},
                "monofasico": is_monofasico_display,
                "aliquota_zero": bool(pis_cofins_zero),
                "nota_aliquota_zero": nota_aliquota_zero,
            },
            "class_trib": class_trib["class"],
            "cst_ipi": class_trib["ipi_cst"],
        },
        "icms_estados": ICMS_STATES,
        "icms_piaui": get_icms_piaui(ncm_clean),
        "icms_hortifruti_aviso": get_icms_hortifruti_aviso(ncm_clean),
        "queijo_aviso": get_queijo_aviso(ncm_clean, descricao_stripped),
        "monofasico": is_monofasico_display,
        "monofasico_dados": mono_data,
        "reforma_tributaria": get_reforma_tributaria_info(ncm_clean, is_monofasico),
    }

    reforma = result["reforma_tributaria"]
    result["classtrib_sugestao"] = get_classtrib_suggestion(
        ncm_clean, is_monofasico, chapter, reforma
    )
    return jsonify(result)


# Incisos da Cesta Básica (Lei 10.925/2004) que NÃO são isenção total para
# fins de LC 214/2025: são insumos agropecuários/veterinários (I-IV, VI,
# VII, X) ou produtos de higiene (XXVI, XXVII), que na tabela cClassTrib
# correspondem a Anexos de redução parcial (Anexo VIII/IX, 60%), não ao
# Anexo I/XV de isenção total (100%).
_REFORMA_ISENCAO_INCISOS_EXCLUIDOS = {"I", "II", "III", "IV", "VI", "VII", "X", "XXVI", "XXVII"}


def _elegivel_isencao_reforma(aliq_zero_info):
    """
    Decide se um resultado de get_pis_cofins_aliquota_zero corresponde a
    isenção total (Anexo I "cesta básica" ou Anexo XV "hortifrúti") para
    fins da LC 214/2025, e não a uma categoria de redução parcial
    (insumos agropecuários, higiene) que usa a mesma alíquota zero de
    PIS/COFINS mas um Anexo diferente na reforma.
    """
    if not aliq_zero_info:
        return False
    modulo = aliq_zero_info.get("modulo")
    if modulo == "Hortifrúti":
        return True
    if modulo == "Cesta Básica":
        return aliq_zero_info.get("inciso") not in _REFORMA_ISENCAO_INCISOS_EXCLUIDOS
    return False


def get_reforma_tributaria_info(ncm_clean, is_monofasico):
    chapter = ncm_clean[:2]
    info = {
        "cbs": {
            "aliquota_referencia": 8.8,
            "descricao": "Contribuição sobre Bens e Serviços (substitui PIS/COFINS)",
            "vigencia": "A partir de 2027 (testes em 2026)",
        },
        "ibs": {
            "aliquota_referencia": 17.7,
            "descricao": "Imposto sobre Bens e Serviços (substitui ICMS/ISS)",
            "vigencia": "Gradual entre 2029-2033",
        },
        "is": {
            "incide": False,
            "descricao": "Imposto Seletivo (sobre produtos nocivos à saúde/meio ambiente)",
            "produtos_sujeitos": [],
        },
        "reducao_aliquota": None,
        "isencao": None,
        "observacoes": [],
    }

    # Combustíveis
    if chapter == "27":
        info["is"]["incide"] = True
        info["is"]["produtos_sujeitos"] = ["Combustíveis fósseis (seletividade ambiental)"]
        info["observacoes"].append("Combustíveis sujeitos ao Imposto Seletivo (LC 214/2025)")
        info["observacoes"].append("Regimes específicos de cobrança monofásica mantidos no CBS/IBS")

    # Tabaco
    elif chapter == "24":
        info["is"]["incide"] = True
        info["is"]["produtos_sujeitos"] = ["Tabaco e derivados"]
        info["observacoes"].append("Produtos de tabaco sujeitos ao Imposto Seletivo")

    # Medicamentos
    elif chapter == "30":
        info["reducao_aliquota"] = 60.0
        info["observacoes"].append("Medicamentos com redução de 60% na alíquota de referência (CBS+IBS)")
        info["observacoes"].append("Medicamentos de uso humano: alíquota efetiva CBS ~3,52% e IBS ~7,08%")
        info["cbs"]["aliquota_efetiva"] = 3.52
        info["ibs"]["aliquota_efetiva"] = 7.08

    # Alimentos básicos / hortifrúti — dirigido pela mesma regra de
    # alíquota zero de PIS/COFINS (Lei 10.925/2004 e 10.865/2004), que já
    # modela com precisão quais NCMs entram (exclui queijo ralado/fundido,
    # leite condensado etc.) em vez de um chapter list genérico, que tanto
    # deixava de fora capítulos inteiros (ex.: açúcar, cap. 17) quanto
    # incluía capítulos inteiros que têm exceções dentro deles (ex.:
    # queijo ralado no cap. 04). Insumos agropecuários (incisos I-IV, VI,
    # VII, X) e higiene (XXVI, XXVII) ficam de fora daqui porque, na
    # tabela cClassTrib real, correspondem a Anexos de REDUÇÃO parcial
    # (Anexo VIII/IX, 60%), não de isenção total (Anexo I/XV).
    elif _elegivel_isencao_reforma(get_pis_cofins_aliquota_zero(ncm_clean)):
        info["isencao"] = "Cesta básica nacional/hortifrúti - isenção total CBS/IBS"
        info["cbs"]["aliquota_efetiva"] = 0.0
        info["ibs"]["aliquota_efetiva"] = 0.0
        info["observacoes"].append("Produto da cesta básica/hortifrúti: alíquota zero CBS e IBS")

    # Veículos
    elif chapter == "87":
        info["is"]["incide"] = True
        info["is"]["produtos_sujeitos"] = ["Veículos automotores (exceto elétricos)"]
        info["observacoes"].append("Veículos com motor de combustão sujeitos ao Imposto Seletivo")
        info["observacoes"].append("Veículos elétricos e híbridos: possível redução de IS")

    # Bebidas alcoólicas
    elif chapter == "22":
        info["is"]["incide"] = True
        info["is"]["produtos_sujeitos"] = ["Bebidas alcoólicas"]
        info["observacoes"].append("Bebidas alcoólicas sujeitas ao Imposto Seletivo")

    if is_monofasico:
        info["observacoes"].append("Produto com tributação monofásica: regime especial mantido na transição")

    return info


def get_classtrib_suggestion(ncm_clean, is_monofasico, chapter, reforma_info):
    """Determina CST-IBS/CBS sugerido e cClassTribs correspondentes para o produto."""

    isencao   = bool(reforma_info.get("isencao"))
    reducao   = reforma_info.get("reducao_aliquota")
    is_incide = reforma_info.get("is", {}).get("incide", False)

    # ---- 1. Isenção (cesta básica / hortifrúti) ----
    # A tabela cClassTrib carregada só tem UMA linha com CST 400 (transporte
    # público), sem nenhuma entrada de isenção de alimentos sob esse CST —
    # por isso NUNCA cravamos um código de 6 dígitos aqui. O mecanismo real
    # de cesta básica/hortifrúti na LC 214/2025 é redução de 100% sob CST
    # 200 (Anexo I / Anexo XV), mas a codificação exata já mudou de versão
    # várias vezes, então exibimos apenas um aviso para o usuário confirmar.
    if isencao:
        return {
            "cst_sugerido": "200/400",
            "desc_cst": "Vendas de produtos destinados à alimentação humana (Anexo I ou XV da LC 214/2025)",
            "motivo": f"Produto com isenção/redução total de IBS e CBS — {reforma_info['isencao']}",
            "base_legal": "Arts. 9º-10 e 120-127 da LC 214/2025 (Cesta Básica Nacional / Hortifrúti)",
            "classtrib_destaque": None,
            "classtrib_list": [],
            "aviso_confirmacao": (
                "Código cClassTrib específico sujeito a confirmação - "
                "consulte a tabela vigente no Portal Nacional da NF-e "
                "(Documentos > Diversos) antes de emitir a nota fiscal."
            ),
        }

    aviso_confirmacao = None

    # 2. Alíquota reduzida (hoje só medicamentos, Capítulo 30) — vem antes
    # de monofásico. pRedCBS na tabela já está em pontos percentuais
    # (60.0 = 60%), não em fração (0.6) — antes a comparação dividia
    # `reducao` por 100, o que nunca batia com nenhuma linha e sempre caía
    # no primeiro item de CST 200 (zonas de processamento de exportação,
    # nada a ver com medicamento). Corrigido para comparar direto, e
    # com filtro por palavra-chave já que várias linhas de CST 200
    # compartilham a mesma redução de 60% para categorias bem diferentes
    # (educação, saúde, medicamentos, agropecuário etc.).
    if reducao:
        destaque_code = None
        if chapter == "30":
            destaque_code = next(
                (r["cClassTrib"] for r in CLASSTRIB_DATA
                 if r["cst_ibs_cbs"] == "200"
                 and abs(r.get("pRedCBS", 0) - reducao) < 0.01
                 and "medicamento" in r["nome"].lower()),
                None
            )
        if destaque_code:
            cst = "200"
            motivo = f"Produto com redução de {reducao:.0f}% nas alíquotas de referência do IBS e CBS."
            base_legal = "Arts. 128-142 da LC 214/2025"
        else:
            # Sem correspondência confiável na tabela carregada — não
            # inventamos um cClassTrib, caímos para tributação integral.
            cst = "000"
            destaque_code = "000001"
            motivo = (
                f"Produto com redução de {reducao:.0f}% nas alíquotas de "
                "referência do IBS e CBS, mas nenhum cClassTrib específico "
                "foi encontrado com confiança na tabela carregada — "
                "exibindo tributação integral como padrão seguro."
            )
            base_legal = "Arts. 1-10 da LC 214/2025"
            aviso_confirmacao = "Confirme o cClassTrib específico na tabela vigente antes de emitir a nota fiscal."

    # 3. Combustíveis monofásicos — único caso com cClassTrib 620 real na
    # tabela carregada (as 6 linhas de CST 620 são todas sobre combustível).
    elif chapter == "27" and is_monofasico:
        cst = "620"
        motivo = "Combustível sujeito à tributação monofásica de IBS e CBS."
        base_legal = "Arts. 154-168 da LC 214/2025"
        destaque_code = "620001"

    # 4. Demais monofásicos no regime atual (bebidas, veículos, autopeças,
    # perfumaria etc.) — a tabela carregada NÃO tem nenhum cClassTrib de
    # monofásico para essas categorias fora de combustíveis (verificado:
    # as 6 linhas de CST 620 mencionam só "combustíveis"). Reaproveitar
    # "620001" aqui mostrava "Tributação monofásica sobre combustíveis"
    # para carro, xampu e cerveja — mesmo tipo de bug do transporte
    # público, só que noutro CST. Caímos para tributação integral.
    elif is_monofasico:
        cst = "000"
        motivo = (
            "Produto com tributação monofásica no regime atual (PIS/COFINS). "
            "A tabela cClassTrib carregada não tem um cClassTrib específico "
            "de monofásico para esta categoria fora de combustíveis — "
            "exibindo tributação integral como padrão seguro."
        )
        base_legal = "Arts. 1-10 da LC 214/2025"
        destaque_code = "000001"
        aviso_confirmacao = "Confirme o tratamento correto na tabela vigente antes de emitir a nota fiscal."

    # 5. Tabaco — integral + IS
    elif chapter == "24":
        cst = "000"
        motivo = "Tributação integral pelo IBS/CBS. Produto sujeito adicionalmente ao Imposto Seletivo."
        base_legal = "Arts. 1-10 da LC 214/2025 (IBS/CBS) + Art. 225 (IS)"
        destaque_code = "000001"

    # 6. Produtos com Imposto Seletivo — tributação integral + IS
    elif is_incide:
        cst = "000"
        motivo = "Tributação integral pelo IBS/CBS. Imposto Seletivo incide adicionalmente sobre este produto."
        base_legal = "Arts. 1-10 da LC 214/2025 (IBS/CBS) + Arts. 225-263 (IS)"
        destaque_code = "000001"

    # 7. Tributação normal
    else:
        cst = "000"
        motivo = "Tributação integral pelo IBS e CBS à alíquota padrão de referência."
        base_legal = "Arts. 1-10 da LC 214/2025"
        destaque_code = "000001"

    # Obtém todos os cClassTrib do CST determinado
    classtrib_for_cst = [r for r in CLASSTRIB_DATA if r["cst_ibs_cbs"] == cst]

    # Garante que destaque_code existe na lista; se não, usa o primeiro
    if not any(r["cClassTrib"] == destaque_code for r in classtrib_for_cst):
        destaque_code = classtrib_for_cst[0]["cClassTrib"] if classtrib_for_cst else None

    desc_cst = next(
        (c["descricao"] for c in CST_IBS_CBS_DATA if c["cst"] == cst), ""
    )

    return {
        "cst_sugerido": cst,
        "desc_cst": desc_cst,
        "motivo": motivo,
        "base_legal": base_legal,
        "classtrib_destaque": destaque_code,
        "classtrib_list": classtrib_for_cst,
        "aviso_confirmacao": aviso_confirmacao,
    }


@app.route('/api/monofasicos')
def get_monofasicos():
    return jsonify(MONOFASICOS)


@app.route('/api/classtrib')
def get_classtrib():
    q = request.args.get('q', '').lower()
    tipo = request.args.get('tipo', '')
    cst = request.args.get('cst', '')
    mono = request.args.get('mono', '')

    result = CLASSTRIB_DATA
    if q:
        result = [r for r in result if
                  q in r['cClassTrib'].lower() or
                  q in r['nome'].lower() or
                  q in r['descricao'].lower() or
                  q in r['cst_ibs_cbs']]
    if tipo:
        result = [r for r in result if r['tipo_aliquota'] == tipo]
    if cst:
        result = [r for r in result if r['cst_ibs_cbs'] == cst]
    if mono == '1':
        result = [r for r in result if r['ind_mono_padrao'] or r['ind_mono_reten'] or r['ind_mono_ret'] or r['ind_mono_dif']]

    return jsonify({
        'total': len(result),
        'data': result,
    })


@app.route('/api/cst-ibs-cbs')
def get_cst_ibs_cbs():
    return jsonify(CST_IBS_CBS_DATA)


def _extrair_ncms_do_texto(texto):
    """
    Extrai NCMs de um texto bruto usando dois padrões distintos.
    Retorna set de strings de exatamente 8 dígitos sem pontuação.

    Regras de validação:
      - Capítulo (2 primeiros dígitos) entre 01 e 99
      - Nunca classifica por nome/descrição — somente pelo código numérico
    """
    encontrados = set()

    # Padrão 1 — formato padrão brasileiro XXXX.XX.XX (pontos obrigatórios)
    # Alta confiança: pontos são exigidos, não opcionais
    for m in re.finditer(r'(?<!\d)(\d{4})\.(\d{2})\.(\d{2})(?!\d)', texto):
        ncm = m.group(1) + m.group(2) + m.group(3)
        if _ncm_capitulo_valido(ncm):
            encontrados.add(ncm)

    # Padrão 2 — 8 dígitos consecutivos isolados (sem dígito antes ou depois)
    # Captura NCMs sem formatação; exige isolamento para evitar falsos positivos
    for m in re.finditer(r'(?<!\d)(\d{8})(?!\d)', texto):
        ncm = m.group(1)
        if _ncm_capitulo_valido(ncm):
            encontrados.add(ncm)

    return encontrados


def _ncm_capitulo_valido(ncm):
    """Valida se os 2 primeiros dígitos formam um capítulo NCM real (01-99)."""
    if len(ncm) != 8:
        return False
    capitulo = int(ncm[:2])
    return 1 <= capitulo <= 99


def _extrair_texto_pdf(pdf_bytes):
    """
    Extrai todo o conteúdo textual do PDF:
    1. Texto corrido de cada página
    2. Células de tabelas (pdfplumber.extract_tables) — captura NCMs em colunas de tabela
    """
    import pdfplumber
    partes = []
    with pdfplumber.open(pdf_bytes) as pdf:
        for page in pdf.pages:
            # Texto corrido
            txt = page.extract_text()
            if txt:
                partes.append(txt)
            # Células de tabelas
            try:
                tabelas = page.extract_tables() or []
                for tabela in tabelas:
                    for linha in (tabela or []):
                        for celula in (linha or []):
                            if celula and isinstance(celula, str):
                                partes.append(celula)
            except Exception:
                pass
    return "\n".join(partes)


@app.route('/api/upload-pdf', methods=['POST'])
def upload_pdf():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({"error": "Apenas arquivos PDF são aceitos"}), 400

    # --- Extrai texto do PDF ---
    try:
        pdf_bytes = io.BytesIO(file.read())
        texto_completo = _extrair_texto_pdf(pdf_bytes)
    except ImportError:
        return jsonify({"error": "Biblioteca pdfplumber não instalada"}), 500
    except Exception as e:
        return jsonify({"error": f"Erro ao processar PDF: {str(e)}"}), 500

    if not texto_completo.strip():
        return jsonify({"error": "Não foi possível extrair texto do PDF. Verifique se o arquivo não é uma imagem escaneada sem OCR."}), 422

    # --- Extrai NCMs do texto (somente por código numérico) ---
    ncms_extraidos = sorted(_extrair_ncms_do_texto(texto_completo))

    if not ncms_extraidos:
        return jsonify({
            "total_ncms": 0,
            "total_monofasicos": 0,
            "total_nao_monofasicos": 0,
            "percentual_monofasico": 0.0,
            "monofasicos": [],
            "nao_monofasicos": [],
            "todos_ncms": [],
            "aviso": "Nenhum NCM encontrado no arquivo. Verifique se o PDF contém códigos no formato XXXX.XX.XX ou 8 dígitos.",
        })

    # --- Cruza NCMs com regras de monofásico (código específico + capítulo) ---
    monofasicos_encontrados = []
    nao_monofasicos = []

    for ncm in ncms_extraidos:
        is_mono, entrada = get_monofasico_info(ncm)
        if is_mono:
            entrada = dict(entrada)
            entrada["ncm_formatado"] = format_ncm(ncm)
            monofasicos_encontrados.append(entrada)
        else:
            nao_monofasicos.append({
                "ncm": ncm,
                "ncm_formatado": format_ncm(ncm),
            })

    total = len(ncms_extraidos)
    percentual = round(len(monofasicos_encontrados) / total * 100, 2) if total else 0.0

    return jsonify({
        "total_ncms": total,
        "total_monofasicos": len(monofasicos_encontrados),
        "total_nao_monofasicos": len(nao_monofasicos),
        "percentual_monofasico": percentual,
        "monofasicos": monofasicos_encontrados,
        "nao_monofasicos": nao_monofasicos,   # agora lista de {ncm, ncm_formatado}
        "todos_ncms": ncms_extraidos,
    })


@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dados não fornecidos"}), 400

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill("solid", fgColor="1A3C6E")
    header_fill_green = PatternFill("solid", fgColor="1B5E3B")
    header_fill_orange = PatternFill("solid", fgColor="B45309")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    green_fill = PatternFill("solid", fgColor="D1FAE5")
    red_fill = PatternFill("solid", fgColor="FEE2E2")

    # --- Aba Resumo ---
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 20

    ws_resumo.merge_cells('A1:B1')
    ws_resumo['A1'] = "RELATÓRIO - ANÁLISE NCM MONOFÁSICOS"
    ws_resumo['A1'].font = Font(bold=True, size=14, color="1A3C6E")
    ws_resumo['A1'].alignment = center
    ws_resumo.row_dimensions[1].height = 30

    resumo_rows = [
        ("Total de NCMs no arquivo", data.get("total_ncms", 0)),
        ("Total de NCMs Monofásicos", data.get("total_monofasicos", 0)),
        ("Total de NCMs Não Monofásicos", data.get("total_nao_monofasicos", 0)),
        ("Percentual Monofásico", f"{data.get('percentual_monofasico', 0):.2f}%"),
    ]
    for i, (label, val) in enumerate(resumo_rows, start=3):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'B{i}'] = val
        ws_resumo[f'A{i}'].font = Font(bold=True)
        ws_resumo[f'B{i}'].alignment = center
        if i % 2 == 0:
            ws_resumo[f'A{i}'].fill = alt_fill
            ws_resumo[f'B{i}'].fill = alt_fill
        ws_resumo[f'A{i}'].border = thin_border
        ws_resumo[f'B{i}'].border = thin_border

    # --- Aba Monofásicos ---
    ws_mono = wb.create_sheet("NCMs Monofásicos")
    ws_mono.column_dimensions['A'].width = 15
    ws_mono.column_dimensions['B'].width = 40
    ws_mono.column_dimensions['C'].width = 22
    ws_mono.column_dimensions['D'].width = 28
    ws_mono.column_dimensions['E'].width = 42

    headers = ["NCM", "Descrição", "Categoria", "Referência Legal", "Regra de Identificação"]
    for col, h in enumerate(headers, 1):
        cell = ws_mono.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = center
        cell.border = thin_border
    ws_mono.row_dimensions[1].height = 25

    for i, item in enumerate(data.get("monofasicos", []), start=2):
        row_data = [
            item.get("ncm_formatado", item.get("ncm", "")),
            item.get("descricao", ""),
            item.get("categoria", ""),
            item.get("referencia", ""),
            item.get("regra_identificacao", "Monofásico por NCM específico"),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_mono.cell(row=i, column=col, value=val)
            cell.alignment = left
            cell.border = thin_border
            if i % 2 == 0:
                cell.fill = alt_fill

    # --- Aba Não Monofásicos ---
    ws_nao = wb.create_sheet("NCMs Não Monofásicos")
    ws_nao.column_dimensions['A'].width = 20
    ws_nao.column_dimensions['B'].width = 50

    cell = ws_nao.cell(row=1, column=1, value="NCM")
    cell.font = header_font
    cell.fill = header_fill_orange
    cell.alignment = center
    cell.border = thin_border

    cell = ws_nao.cell(row=1, column=2, value="Observação")
    cell.font = header_font
    cell.fill = header_fill_orange
    cell.alignment = center
    cell.border = thin_border
    ws_nao.row_dimensions[1].height = 25

    for i, ncm in enumerate(data.get("nao_monofasicos", []), start=2):
        ws_nao.cell(row=i, column=1, value=ncm).border = thin_border
        ws_nao.cell(row=i, column=2, value="Tributação Normal (não monofásico)").border = thin_border
        if i % 2 == 0:
            ws_nao.cell(row=i, column=1).fill = alt_fill
            ws_nao.cell(row=i, column=2).fill = alt_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='relatorio_ncm_monofasicos.xlsx'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port)
